"""Laboratorio del bloque 16: SQLite -> controles -> libro Excel.

Instala dependencias: pip install -r notebooks/practicas/requirements-bloque-16.txt
Los importes se almacenan y concilian como céntimos enteros. Excel recibe euros
solo para lectura; no se utiliza coma flotante para los controles monetarios.
"""

from __future__ import annotations

import argparse
import logging
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[2]
DATABASE = ROOT / "datasets" / "temario-16" / "operaciones.sqlite"
OUTPUT = ROOT / "salidas"
REQUIRED_COLUMNS = {"operacion_id", "fecha_utc", "estado", "importe_centimos", "canal", "es_prueba"}
ESTADOS_NO_PAGADOS = {
    "rechazada": "Cobro no autorizado o fallido",
    "pendiente": "Resultado aún no definitivo",
    "devuelta": "Pago revertido posteriormente",
}


def crear_base_si_falta() -> None:
    """Crea una fuente didáctica; no modifica una fuente empresarial existente."""
    DATABASE.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DATABASE) as conexion:
        columnas = {fila[1] for fila in conexion.execute("PRAGMA table_info(operaciones)")}
        if columnas and "importe_centimos" not in columnas:
            # Migración solo para la base didáctica que usaba importe_eur REAL.
            conexion.execute("ALTER TABLE operaciones ADD COLUMN importe_centimos INTEGER")
            conexion.execute("UPDATE operaciones SET importe_centimos = ROUND(importe_eur * 100)")
        conexion.execute(
            """CREATE TABLE IF NOT EXISTS operaciones (
                operacion_id TEXT PRIMARY KEY, fecha_utc TEXT NOT NULL,
                estado TEXT NOT NULL, importe_centimos INTEGER, canal TEXT NOT NULL,
                es_prueba INTEGER NOT NULL DEFAULT 0
            )"""
        )
        if conexion.execute("SELECT COUNT(*) FROM operaciones").fetchone()[0] == 0:
            conexion.executemany(
                "INSERT INTO operaciones VALUES (?, ?, ?, ?, ?, ?)",
                [
                    ("op-001", "2026-07-13T08:00:00Z", "pagada", 12000, "web", 0),
                    ("op-002", "2026-07-14T09:30:00Z", "rechazada", 8000, "app", 0),
                    ("op-003", "2026-07-15T10:00:00Z", "pagada", 5000, "web", 0),
                    ("op-004", "2026-07-16T14:15:00Z", "devuelta", 3000, "partner", 0),
                    ("op-005", "2026-07-17T12:20:00Z", "pendiente", 2500, "app", 0),
                    ("op-test", "2026-07-18T10:00:00Z", "pagada", 100, "web", 1),
                ],
            )


def cargar(inicio: str, fin: str) -> pd.DataFrame:
    sql = """
    SELECT operacion_id, fecha_utc, estado, importe_centimos, canal, es_prueba
    FROM operaciones
    WHERE fecha_utc >= :inicio AND fecha_utc < :fin
    """
    with sqlite3.connect(DATABASE) as conexion:
        return pd.read_sql_query(sql, conexion, params={"inicio": inicio, "fin": fin})


def total_pagado_sql(inicio: str, fin: str) -> int:
    """Control independiente: el total se calcula de nuevo en la fuente SQL."""
    sql = """
    SELECT COALESCE(SUM(importe_centimos), 0)
    FROM operaciones
    WHERE fecha_utc >= :inicio AND fecha_utc < :fin
      AND es_prueba = 0 AND estado = 'pagada'
    """
    with sqlite3.connect(DATABASE) as conexion:
        return int(conexion.execute(sql, {"inicio": inicio, "fin": fin}).fetchone()[0])


def fila_control(control: str, valor: object, esperado: object, superado: bool, accion: str) -> dict[str, object]:
    return {"control": control, "valor": valor, "esperado": esperado, "superado": superado, "accion_si_falla": accion}


def validar(datos: pd.DataFrame, inicio: str, fin: str, total_sql: int | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Devuelve datos tipados y controles; un esquema inválido nunca causa KeyError."""
    faltantes = REQUIRED_COLUMNS - set(datos.columns)
    if faltantes:
        controles = pd.DataFrame([fila_control(
            "columnas requeridas", ", ".join(sorted(faltantes)), "ninguna columna faltante", False,
            "Bloquear la entrega y corregir el esquema de origen.",
        )])
        return datos.copy(), controles

    preparados = datos.copy()
    preparados["fecha_utc"] = pd.to_datetime(preparados["fecha_utc"], utc=True, errors="coerce")
    preparados["importe_centimos"] = pd.to_numeric(preparados["importe_centimos"], errors="coerce")
    inicio_dt, fin_dt = pd.Timestamp(inicio, tz="UTC"), pd.Timestamp(fin, tz="UTC")
    elegibles = preparados.query("es_prueba == 0")
    pagadas = elegibles.query("estado == 'pagada'")
    no_pagadas = elegibles.query("estado != 'pagada'")
    conteos_estado = elegibles["estado"].value_counts().to_dict()
    identidad = len(elegibles) == len(pagadas) + len(no_pagadas)
    total_dataframe = int(pagadas["importe_centimos"].sum()) if pagadas["importe_centimos"].notna().all() else None
    total_sql = total_dataframe if total_sql is None else total_sql
    controles = [
        fila_control("columnas requeridas", len(REQUIRED_COLUMNS), len(REQUIRED_COLUMNS), True, "No aplica."),
        fila_control("filas extraídas", len(preparados), ">= 1", len(preparados) >= 1, "Confirmar periodo o carga de origen."),
        fila_control("IDs únicos", int(preparados["operacion_id"].duplicated().sum()), 0, not preparados["operacion_id"].duplicated().any(), "Bloquear y revisar grano o reintentos."),
        fila_control("fechas dentro de [inicio, fin)", f"{preparados['fecha_utc'].min()} — {preparados['fecha_utc'].max()}", f"[{inicio}, {fin})", preparados["fecha_utc"].notna().all() and preparados["fecha_utc"].ge(inicio_dt).all() and preparados["fecha_utc"].lt(fin_dt).all(), "Bloquear y revisar zona horaria o filtros."),
        fila_control("importe pagado no nulo", int(pagadas["importe_centimos"].isna().sum()), 0, pagadas["importe_centimos"].notna().all(), "Bloquear y corregir importes de pagos."),
        fila_control("filas de prueba excluidas", int((preparados["es_prueba"] == 1).sum()), "registradas, no incluidas", True, "Investigar si una prueba entra en elegibles."),
        fila_control("identidad elegibles = pagadas + no pagadas", len(elegibles), f"{len(pagadas)} + {len(no_pagadas)}", identidad, "Bloquear y revisar clasificación de estados."),
        fila_control("total pagado DataFrame (céntimos)", total_dataframe, total_sql, total_dataframe == total_sql, "Bloquear y conciliar con la fuente SQL."),
    ]
    for estado, cantidad in sorted(conteos_estado.items()):
        controles.append(fila_control(f"estado: {estado}", cantidad, "informativo", True, "Revisar cambios inesperados."))
    return preparados, pd.DataFrame(controles)


def preparar_para_excel(datos: pd.DataFrame) -> pd.DataFrame:
    salida = datos.copy()
    salida["fecha_utc"] = salida["fecha_utc"].dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    salida["importe_eur"] = salida["importe_centimos"] / 100
    return salida.drop(columns="importe_centimos")


def aplicar_tabla(hoja, nombre: str, referencia: str) -> None:
    tabla = Table(displayName=nombre, ref=referencia)
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    hoja.add_table(tabla)


def formatear_hoja(hoja, encabezados: list[int], monetarias: list[str] = ()) -> None:
    azul, blanco = "1D5D84", "FFFFFF"
    for fila in encabezados:
        for celda in hoja[fila]:
            if celda.value is not None:
                celda.font = Font(bold=True, color=blanco)
                celda.fill = PatternFill("solid", fgColor=azul)
    for columna in monetarias:
        for celda in hoja[columna][1:]:
            celda.number_format = '#,##0.00 [$€-es-ES]'
    for columna in hoja.columns:
        letra = columna[0].column_letter
        hoja.column_dimensions[letra].width = min(44, max(12, max(len(str(c.value or "")) for c in columna) + 2))
    hoja.freeze_panes = "A5" if hoja.title == "Resumen" else "A2"
    hoja.sheet_view.showGridLines = False


def generar_libro(datos: pd.DataFrame, controles: pd.DataFrame, inicio: str, fin: str) -> Path:
    salida = OUTPUT / f"operaciones_{inicio}_a_{fin}.xlsx"
    OUTPUT.mkdir(exist_ok=True)
    elegibles = datos.query("es_prueba == 0").copy()
    pagadas = elegibles.query("estado == 'pagada'").copy()
    no_pagadas = elegibles.query("estado != 'pagada'").copy()
    no_pagadas["motivo_exclusion_total"] = no_pagadas["estado"].map(ESTADOS_NO_PAGADOS).fillna("Estado no reconocido: investigar antes de entregar")
    resumen = pd.DataFrame({"indicador": ["Intentos elegibles", "Operaciones pagadas", "Importe cobrado EUR"], "valor": [len(elegibles), len(pagadas), int(pagadas["importe_centimos"].sum()) / 100]})
    por_canal = pagadas.groupby("canal", as_index=False)["importe_centimos"].sum()
    por_canal["importe_pagado_eur"] = por_canal.pop("importe_centimos") / 100
    metadatos = pd.DataFrame({"campo": ["inicio_inclusivo_utc", "fin_exclusivo_utc", "generado_utc", "fuente", "estado_controles", "instruccion"], "valor": [inicio, fin, datetime.now(timezone.utc).isoformat(), "SQLite didáctica: operaciones", "APTO" if controles["superado"].all() else "BLOQUEADO", "Leer Resumen; si algún control falla, no comunicar el total."]})
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen", index=False, startrow=3)
        por_canal.to_excel(writer, sheet_name="Resumen", index=False, startrow=8)
        preparar_para_excel(elegibles).to_excel(writer, sheet_name="Detalle", index=False)
        preparar_para_excel(no_pagadas).to_excel(writer, sheet_name="No_pagadas", index=False)
        controles.to_excel(writer, sheet_name="Conciliacion", index=False)
        metadatos.to_excel(writer, sheet_name="Metadatos", index=False)
    libro = load_workbook(salida)
    resumen_ws = libro["Resumen"]
    resumen_ws["A1"] = "Informe semanal de operaciones"
    resumen_ws["A2"] = f"Periodo UTC: [{inicio}, {fin}) · Revisión: consulte Conciliacion antes de comunicar cifras."
    resumen_ws["A1"].font = Font(bold=True, size=16, color="12355B")
    resumen_ws["A2"].alignment = Alignment(wrap_text=True)
    aplicar_tabla(resumen_ws, "ResumenIndicadores", "A4:B7")
    aplicar_tabla(resumen_ws, "ResumenCanales", f"A9:B{9 + len(por_canal)}")
    aplicar_tabla(libro["Detalle"], "DetalleOperaciones", f"A1:F{len(elegibles) + 1}")
    aplicar_tabla(libro["No_pagadas"], "NoPagadas", f"A1:G{len(no_pagadas) + 1}")
    aplicar_tabla(libro["Conciliacion"], "ControlesInforme", f"A1:E{len(controles) + 1}")
    aplicar_tabla(libro["Metadatos"], "MetadatosInforme", f"A1:B{len(metadatos) + 1}")
    formatear_hoja(resumen_ws, [4, 9])
    resumen_ws["B7"].number_format = '#,##0.00 [$€-es-ES]'
    for fila in range(10, 10 + len(por_canal)):
        resumen_ws.cell(fila, 2).number_format = '#,##0.00 [$€-es-ES]'
    formatear_hoja(libro["Detalle"], [1], ["F"])
    formatear_hoja(libro["No_pagadas"], [1], ["G"])
    formatear_hoja(libro["Conciliacion"], [1])
    formatear_hoja(libro["Metadatos"], [1])
    controles_ws = libro["Conciliacion"]
    controles_ws.conditional_formatting.add(f"D2:D{len(controles) + 1}", CellIsRule(operator="equal", formula=["FALSE"], fill=PatternFill("solid", fgColor="FECACA")))
    libro["Metadatos"].protection.sheet = True  # Evita cambios accidentales; no sustituye permisos reales.
    libro.save(salida)
    return salida


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inicio", default="2026-07-13")
    parser.add_argument("--fin", default="2026-07-20")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    crear_base_si_falta()
    datos = cargar(args.inicio, args.fin)
    datos, controles = validar(datos, args.inicio, args.fin, total_pagado_sql(args.inicio, args.fin))
    if not controles["superado"].all():
        logging.error("Informe bloqueado:\n%s", controles.to_string(index=False))
        raise SystemExit(2)
    ruta = generar_libro(datos, controles, args.inicio, args.fin)
    logging.info("Informe generado: %s", ruta)


if __name__ == "__main__":
    main()
