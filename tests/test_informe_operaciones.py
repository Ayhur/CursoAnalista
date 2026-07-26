"""Pruebas de regresión del laboratorio del bloque 16."""

from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "notebooks" / "practicas" / "16-informe-operaciones.py"
SPEC = importlib.util.spec_from_file_location("informe_operaciones", SCRIPT)
modulo = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
SPEC.loader.exec_module(modulo)


class InformeOperacionesTest(unittest.TestCase):
    def test_esquema_invalido_devuelve_control_y_no_keyerror(self) -> None:
        datos, controles = modulo.validar(pd.DataFrame({"fecha_utc": []}), "2026-07-13", "2026-07-20")
        self.assertFalse(controles.loc[0, "superado"])
        self.assertIn("operacion_id", controles.loc[0, "valor"])
        self.assertEqual(list(datos.columns), ["fecha_utc"])

    def test_controles_concilian_cientos_y_clasificacion(self) -> None:
        datos = pd.DataFrame(
            [
                ("a", "2026-07-13T08:00:00Z", "pagada", 12000, "web", 0),
                ("b", "2026-07-13T09:00:00Z", "devuelta", 3000, "web", 0),
                ("c", "2026-07-13T10:00:00Z", "pendiente", 2000, "app", 0),
                ("d", "2026-07-13T11:00:00Z", "pagada", 100, "web", 1),
            ],
            columns=["operacion_id", "fecha_utc", "estado", "importe_centimos", "canal", "es_prueba"],
        )
        preparados, controles = modulo.validar(datos, "2026-07-13", "2026-07-20", total_sql=12000)
        self.assertTrue(controles["superado"].all())
        self.assertEqual(preparados.loc[0, "importe_centimos"], 12000)
        self.assertIn("identidad elegibles", " ".join(controles["control"]))

    def test_libro_contiene_hojas_tablas_y_motivos(self) -> None:
        with tempfile.TemporaryDirectory() as temporal:
            modulo.OUTPUT = Path(temporal)
            datos = pd.DataFrame(
                [
                    ("a", "2026-07-13T08:00:00Z", "pagada", 12000, "web", 0),
                    ("b", "2026-07-13T09:00:00Z", "devuelta", 3000, "web", 0),
                ],
                columns=["operacion_id", "fecha_utc", "estado", "importe_centimos", "canal", "es_prueba"],
            )
            preparados, controles = modulo.validar(datos, "2026-07-13", "2026-07-20", total_sql=12000)
            salida = modulo.generar_libro(preparados, controles, "2026-07-13", "2026-07-20")
            libro = load_workbook(salida, read_only=False)
            self.assertEqual(libro.sheetnames, ["Resumen", "Detalle", "No_pagadas", "Conciliacion", "Metadatos"])
            self.assertIn("NoPagadas", libro["No_pagadas"].tables)
            encabezados = {celda.value: celda.column for celda in libro["No_pagadas"][1]}
            self.assertEqual(
                libro["No_pagadas"].cell(2, encabezados["motivo_exclusion_total"]).value,
                "Pago revertido posteriormente",
            )


if __name__ == "__main__":
    unittest.main()
